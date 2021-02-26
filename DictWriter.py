import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import re
import time
from threading import Thread


def find_smaller(a, number):
    l_index = len(a) - 1
    for i in range(len(a)):
        if i == l_index:
            break
        elif a[i] < number & number < a[i+1]:
            l_index = i
            break
    return l_index


class DictWriter:
    def __init__(self, word, sheet):
        self.mark = word
        self.name = word.lstrip('-')
        self.pos = 0
        self.soup = None
        self.text = None
        self.means = []
        self.meanCounts = []
        self.examples = []
        self.examCounts = []
        self.props = []
        self.propCounts = []
        self.ipa = None
        self.st_s = 0
        self.st_e = 0
        self.ot_s = 0
        self.ot_e = 0
        self.sheet = sheet

    def __str__(self):
        return self.name

    def get_soup_and_text(self):
        res = requests.get(f'https://dictionary.cambridge.org/us/dictionary/english/{self.name}',
                           headers={"User-Agent": "Mozilla/5.0"})
        soup_str = res.text
        pattern_to_cut = re.compile(r'Intermediate English')
        matches_to_cut = pattern_to_cut.finditer(soup_str)

        for match in matches_to_cut:
            end = match.start()
            soup_str = soup_str[:end]
            break

        self.soup = BeautifulSoup(soup_str, 'html.parser')
        fulltext = self.soup.getText()
        textures = re.sub(r'^ *', r'', fulltext, flags=re.M)
        self.text = re.sub(r'(\n){1,1000}', r'\n', textures, flags=re.M)

    def get_meaning_starts(self):
        mean_tab = self.soup.find_all(class_='def ddef_d db')
        for i in mean_tab:
            self.means.append(i.getText())
        for m in self.means:
            a = re.escape(m)
            pattern = re.compile(a)
            matches = pattern.finditer(self.text)
            count = 0
            for match in matches:
                if count:
                    self.means.remove(m)
                    break
                self.meanCounts.append(match.start())
                count += 1

    def get_exam_starts(self):
        eg_tab = self.soup.find_all(class_='eg deg')
        for i in eg_tab:
            self.examples.append(i.getText())
        for m in self.examples:
            a = re.escape(m)
            pattern = re.compile(a)
            matches = pattern.finditer(self.text)
            count = 0
            for match in matches:
                if count:
                    self.examples.remove(m)
                    break
                self.examCounts.append(match.start())
                count += 1

    def get_prop_starts(self):
        pattern = re.compile(r'^' + self.name + r'( ?(\w{3,7})\D{0,30})us', flags=re.M)
        matches = pattern.finditer(self.text)
        for match in matches:
            text_to = match.group()
            text_final = re.sub(r'(' + self.name + r')' + r'|(us)', '', text_to)

            index = 0
            for i in range(len(text_final)):
                if text_final[i] == '\n':
                    index = i
            text_final = text_final[index:-1] + text_final[-1]

            self.props.append(text_final.lstrip())
            self.propCounts.append(match.start())

    def get_ipa_title(self):
        ipa_tab = self.soup.find_all(class_='ipa dipa lpr-2 lpl-1')
        for i in ipa_tab:
            self.ipa = f'/ {i.getText()} /'
            break

    def get_contents(self):
        s_t = time.time()
        self.get_soup_and_text()
        self.get_meaning_starts()
        self.get_exam_starts()
        self.get_prop_starts()
        self.get_ipa_title()
        e_t = time.time()
        print(f'{self.name}{self.ipa} takes {(e_t-s_t):.2f} seconds!')

    def show_contents(self):

        a = f'!!False {len(self.means)}-{len(self.meanCounts)}' if len(self.means) != len(self.meanCounts) else 1
        b = f'!!False {len(self.examples)}-{len(self.examCounts)}' if len(self.examples) != len(self.examCounts) else 1
        c = f'!!False {len(self.props)}-{len(self.propCounts)}' if len(self.props) != len(self.propCounts) else 1
        print(self.name)
        print(a, self.means)
        print(b, self.examples)
        print(c, self.props)
        print(self.ipa)
        print('-------------------------')

    # Fill in contents
    def set_pos(self, pos):
        self.pos = pos

    def cell_style(self, number):
        is_bold = 1 if self.mark[0] != '-' else 0
        is_size = 11 if self.mark[0] != '-' else 9
        fonts = [Font(name='Open sans', size=is_size, bold=is_bold),  # title
                 Font(name='Open sans', size=6, bold=0),  # meaning and example
                 ]
        alignment1 = Alignment(vertical='center', wrap_text=1)
        alignment2 = Alignment(horizontal='right', vertical='center', wrap_text=1)
        if number == 3:
            return alignment1
        elif number == 4:
            return alignment2
        else:
            return fonts[number-1]

    def combine_props(self):
        for m_num, m in enumerate(self.means):
            index = find_smaller(self.propCounts, self.meanCounts[m_num])
            prop = self.props[index]
            self.means[m_num] = f'{prop}, ( ) → {self.means[m_num]} '

    def fill_means(self):
        for m_num, m in enumerate(self.means):
            c = self.sheet.cell(row=m_num+self.pos, column=2)
            c.value = m
            c.font = self.cell_style(2)
            c.alignment = self.cell_style(3)

    def fill_exams(self):
        for m_num in range(len(self.means)):
            c = self.sheet.cell(row=m_num+self.pos, column=3)
            c.value = ''
            values = []
            for i in range(len(self.examples)):
                index = find_smaller(self.meanCounts, self.examCounts[i])
                if index == m_num:
                    values.append(f'{self.examples[i]}')
                else:
                    pass

            if len(values) > 3:
                values = values[1:4]  # strip examples to 3.

            for i in range(1, len(values)+1):
                if len(values) != 0:
                    tail = '\n' if i != len(values) else ''
                    c.value += f'{i}. {values[i-1]}{tail}'
                    c.font = self.cell_style(2)
                    c.alignment = self.cell_style(3)
                else:
                    print('No example for this meaning')

    def fill_title(self):
        c = self.sheet.cell(row=self.pos, column=1)
        c.value = f'{self.name}\n{self.ipa}'
        c.font = self.cell_style(1)

        alig = 3
        if self.mark[0] == '-':
            alig = 4
        c.alignment = self.cell_style(alig)
        self.sheet.merge_cells(start_row=self.pos, start_column=1, end_row=self.pos+len(self.means)-1,
                               end_column=1)

    def fill_contents(self):
        self.combine_props()
        self.fill_means()
        self.fill_exams()
        self.fill_title()


def main():
    # Initial
    wb = load_workbook(r'C:\Users\wangshu\Desktop\English Learning\temp.xlsx')
    ws = wb.active
    with open(r'C:\Users\wangshu\Desktop\English Learning\words.txt', 'r', encoding='utf-8') as f:
        words = f.readlines()
    cases = []
    for word in words:
        cases.append(DictWriter(word.rstrip(), ws))

    lost_word = {}

    # Get contents
    threads = []
    for case in cases:
        thread = Thread(target=case.get_contents)
        threads.append(thread)
        thread.start()
    for i, thread in enumerate(threads):
        try:
            thread.join()
        except requests.exceptions.ProxyError:
            lost_word[cases[i].name] = 'Proxy Error'

    # Filling words
    positions = [1]
    for case in cases:
        positions.append(positions[-1]+len(case.means))

    for i, case in enumerate(cases):
        # print(f'{i+1}. {case.name}')
        if case.ipa is None:
            lost_word[case.name] = "Word Not Found"
            continue
        case.set_pos(positions[i])
        case.fill_contents()

    print('-'*50)
    lost_word_count = 0
    for key, value in lost_word.items():
        lost_word_count += 1
        print(f"Lost Words:\n{lost_word_count}. {key}: {value}")
    wb.save(r'C:\Users\wangshu\Desktop\English Learning\temp.xlsx')


if __name__ == '__main__':
    main()
